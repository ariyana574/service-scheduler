from pathlib import Path
import math
import shutil
import re
import pandas as pd
from math import radians, sin, cos, asin, sqrt
from typing import Optional

from openpyxl import load_workbook


DEFAULT_MONTHLY_CAP = 45

BASE_DIR = Path(__file__).resolve().parent.parent
_DEFAULT_INPUT_DIR = BASE_DIR / "data" / "input"
_DEFAULT_OUTPUT_DIR = BASE_DIR / "data" / "output"


# Postcode handling

def detect_postcode_column(df: pd.DataFrame) -> str:
    # Different exports name the postcode column differently, so try the names
    # we've seen first, then fall back to anything containing "post".
    possible_names = [
        "postcode",
        "site postal code",
        "site_postal_code",
        "post code",
        "sitepostcode",
        "site_post_code",
    ]

    lower_map = {c.lower(): c for c in df.columns}

    for known in possible_names:
        if known in lower_map:
            return lower_map[known]

    for col in df.columns:
        if "post" in col.lower():
            return col

    raise ValueError(
        f"Could not detect a postcode column. Columns found: {df.columns.tolist()}"
    )


def add_postal_area_to_jobs(jobs: pd.DataFrame) -> pd.DataFrame:
    jobs = jobs.copy()
    postcode_col = detect_postcode_column(jobs)
    print(f"[jobs] Using '{postcode_col}' as the postcode column in job_data.xlsx")

    jobs[postcode_col] = jobs[postcode_col].astype(str).str.strip()
    # Outcode is the first half of the postcode (e.g. B24); area is just the letters (B).
    jobs["Outcode"] = jobs[postcode_col].str.extract(r"^([A-Z]{1,2}\d{1,2})", expand=False)
    jobs["postalArea"] = jobs[postcode_col].str.extract(r"^([A-Z]{1,2})", expand=False)
    return jobs


def add_postal_area_to_engineers(engineers: pd.DataFrame) -> pd.DataFrame:
    engineers = engineers.copy()
    postcode_col = detect_postcode_column(engineers)
    print(f"[engineers] Using '{postcode_col}' as the postcode column in engineers.xlsx")

    engineers[postcode_col] = engineers[postcode_col].astype(str).str.strip().str.upper()
    engineers["Outcode"] = engineers[postcode_col].str.extract(r"^([A-Z]{1,2}\d{1,2})", expand=False)
    engineers["postalArea"] = engineers[postcode_col].str.extract(r"^([A-Z]{1,2})", expand=False)
    return engineers


# Distances and centroids

def haversine_km(lat1, lon1, lat2, lon2) -> float:
    # Straight-line distance between two lat/lon points in km. Missing coords
    # come back as infinity so they sort to the bottom when picking nearest.
    if any(pd.isna(x) for x in [lat1, lon1, lat2, lon2]):
        return float("inf")

    lat1, lon1, lat2, lon2 = map(radians, [lat1, lon1, lat2, lon2])
    dlat = lat2 - lat1
    dlon = lon2 - lon1
    a = sin(dlat / 2) ** 2 + cos(lat1) * cos(lat2) * sin(dlon / 2) ** 2
    c = 2 * asin(sqrt(a))
    return c * 6371


def build_geo_maps(outcodes: pd.DataFrame):
    oc = outcodes.copy()
    oc["postcode"] = oc["postcode"].astype(str).str.strip().str.upper()
    oc["Outcode"] = oc["postcode"].str.extract(r"^([A-Z]{1,2}\d{1,2})", expand=False)
    oc["postalArea"] = oc["postcode"].str.extract(r"^([A-Z]{1,2})", expand=False)

    # Average the points in each outcode and each area to get a single coordinate.
    outcode_coords = (
        oc.dropna(subset=["Outcode"])
          .groupby("Outcode")[["latitude", "longitude"]]
          .mean()
          .dropna()
          .to_dict("index")
    )

    area_centroids = (
        oc.dropna(subset=["postalArea"])
          .groupby("postalArea")[["latitude", "longitude"]]
          .mean()
          .dropna()
          .to_dict("index")
    )

    return outcode_coords, area_centroids


# Turning the master schedule into 6-month visit cycles

MONTH_ORDER = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_TO_NUM = {m: i + 1 for i, m in enumerate(MONTH_ORDER)}
NUM_TO_MONTH = {v: k for k, v in MONTH_TO_NUM.items()}


def build_schedule_cycles(schedule: pd.DataFrame) -> pd.DataFrame:
    sched = schedule.copy()

    if "Postcode" not in sched.columns:
        raise ValueError("Expected a 'Postcode' column in SERVICE_SCHEDULE_MASTER.xlsx")

    sched["postalArea"] = sched["Postcode"].astype(str).str.extract(r"^([A-Z]{1,2})", expand=False)

    records = []
    for _, row in sched.iterrows():
        postcode = str(row["Postcode"]).strip()
        postal_area = row["postalArea"]

        # A month counts as active if it has any value in it.
        active_months = []
        for m in MONTH_ORDER:
            if m in row.index:
                val = row[m]
                if pd.notna(val) and str(val).strip() not in ("", "0"):
                    active_months.append(MONTH_TO_NUM[m])

        if not active_months:
            continue

        active_months = sorted(active_months)

        # A single active month is just one annual visit.
        if len(active_months) == 1:
            m = active_months[0]
            records.append({
                "Postcode": postcode,
                "postalArea": postal_area,
                "CycleID": 1,
                "Month1": NUM_TO_MONTH[m],
                "Month2": None,
                "IsAnnualOnly": True,
            })
            continue

        # Otherwise pair months that sit 6 months apart into one cycle.
        used = set()
        cycle_id = 1

        for m1 in active_months:
            if m1 in used:
                continue
            partner = None
            for m2 in active_months:
                if m2 in used or m2 == m1:
                    continue
                if (m2 - m1) % 12 == 6:
                    partner = m2
                    break

            if partner is not None:
                records.append({
                    "Postcode": postcode,
                    "postalArea": postal_area,
                    "CycleID": cycle_id,
                    "Month1": NUM_TO_MONTH[m1],
                    "Month2": NUM_TO_MONTH[partner],
                    "IsAnnualOnly": False,
                })
                used.add(m1)
                used.add(partner)
                cycle_id += 1
            else:
                # No 6-month partner, so this month stands on its own.
                records.append({
                    "Postcode": postcode,
                    "postalArea": postal_area,
                    "CycleID": cycle_id,
                    "Month1": NUM_TO_MONTH[m1],
                    "Month2": None,
                    "IsAnnualOnly": False,
                })
                used.add(m1)
                cycle_id += 1

    return pd.DataFrame(records)


# Ordering engineers by distance to each area

def build_area_engineer_distance_map(
    engineers_with_area: pd.DataFrame,
    outcode_coords: dict,
    area_centroids: dict,
    cycles_df: pd.DataFrame,
):
    eng = engineers_with_area.copy()

    # Build a display name from whatever name columns are present.
    first_col = next((c for c in eng.columns if c.lower() in ("first name", "firstname", "first_name")), None)
    last_col = next((c for c in eng.columns if c.lower() in ("last name", "lastname", "last_name")), None)

    if first_col and last_col:
        eng["EngineerName"] = eng[first_col].astype(str).str.strip() + " " + eng[last_col].astype(str).str.strip()
    elif "Employee Id" in eng.columns:
        eng["EngineerName"] = eng["Employee Id"].astype(str)
    else:
        eng["EngineerName"] = eng.index.astype(str)

    # Give every engineer a coordinate: prefer their exact outcode, fall back to
    # the area centroid if we don't have the outcode.
    eng["lat"] = None
    eng["lon"] = None

    for idx, row in eng.iterrows():
        outcode = row.get("Outcode")
        area = row.get("postalArea")

        lat = lon = None
        if isinstance(outcode, str) and outcode in outcode_coords:
            lat = outcode_coords[outcode]["latitude"]
            lon = outcode_coords[outcode]["longitude"]
        elif isinstance(area, str) and area in area_centroids:
            lat = area_centroids[area]["latitude"]
            lon = area_centroids[area]["longitude"]

        eng.at[idx, "lat"] = lat
        eng.at[idx, "lon"] = lon

    eng_valid = eng.dropna(subset=["lat", "lon"])
    all_engineers = eng_valid["EngineerName"].tolist()
    area_eng_map = {}

    # For each area, sort engineers nearest-first.
    for area in cycles_df["postalArea"].dropna().unique():
        if area not in area_centroids:
            area_eng_map[area] = all_engineers[:]
            continue

        area_lat = area_centroids[area]["latitude"]
        area_lon = area_centroids[area]["longitude"]

        distances = [
            (erow["EngineerName"], haversine_km(area_lat, area_lon, erow["lat"], erow["lon"]))
            for _, erow in eng_valid.iterrows()
        ]
        distances_sorted = sorted(distances, key=lambda x: x[1])
        area_eng_map[area] = [name for name, _ in distances_sorted]

    return area_eng_map, all_engineers


# Spreading jobs across the year and assigning engineers

def assign_jobs_to_engineers(cycles_df, jobs_with_area, area_eng_map,
                             all_engineers, monthly_cap):
    jobs_counts = jobs_with_area.groupby("postalArea").size().to_dict()

    # Work out how many visits each area gets, so we can split its jobs evenly.
    visits_per_area = {}
    for area, group in cycles_df.groupby("postalArea"):
        visits = sum(2 if pd.notna(row["Month2"]) and row["Month2"] is not None else 1
                     for _, row in group.iterrows())
        visits_per_area[area] = max(visits, 1)

    engineer_load = {}
    assignment_records = []
    base_year = 2025

    for area, group in cycles_df.groupby("postalArea"):
        total_jobs = jobs_counts.get(area, 0)
        if total_jobs == 0:
            continue

        jobs_per_visit = math.ceil(total_jobs / visits_per_area.get(area, 1))
        candidate_engineers = area_eng_map.get(area) or all_engineers

        if not candidate_engineers:
            continue

        primary_engineer = candidate_engineers[0]

        for _, row in group.iterrows():
            months = []
            m1, m2 = row["Month1"], row["Month2"]

            if isinstance(m1, str):
                months.append(MONTH_TO_NUM[m1])
            elif isinstance(m1, int):
                months.append(m1)

            if pd.notna(m2):
                if isinstance(m2, str):
                    months.append(MONTH_TO_NUM[m2])
                elif isinstance(m2, int):
                    months.append(m2)

            for m in months:
                year_month = f"{base_year}-{m:02d}"
                assigned_engineer = None
                spillover = False
                over_cap = False

                # Walk engineers nearest-first and give the work to the first one
                # who still has room that month.
                for idx, eng_name in enumerate(candidate_engineers):
                    key = (eng_name, year_month)
                    current_load = engineer_load.get(key, 0)
                    if current_load + jobs_per_visit <= monthly_cap:
                        assigned_engineer = eng_name
                        engineer_load[key] = current_load + jobs_per_visit
                        if idx > 0:
                            # Nearest engineer was full, so this spilled over to another.
                            spillover = True
                        break

                # Everyone was full: hand it to the nearest engineer anyway and
                # flag it as over capacity.
                if assigned_engineer is None:
                    key = (primary_engineer, year_month)
                    engineer_load[key] = engineer_load.get(key, 0) + jobs_per_visit
                    assigned_engineer = primary_engineer
                    over_cap = True

                assignment_records.append({
                    "YearMonth": year_month,
                    "postalArea": area,
                    "PostcodeExample": row["Postcode"],
                    "CycleID": row["CycleID"],
                    "PlannedMonthName": NUM_TO_MONTH[m],
                    "Jobs_Assigned": jobs_per_visit,
                    "Engineer": assigned_engineer,
                    "IsSpillover": spillover,
                    "IsOverCap": over_cap,
                })

    assignments_df = pd.DataFrame(assignment_records)

    summary_records = [
        {
            "Engineer": eng_name,
            "YearMonth": year_month,
            "Jobs_Assigned": count,
            "Capacity": monthly_cap,
            "UtilisationPct": round(100 * count / monthly_cap, 1),
            "Status": (
                "Over cap" if count > monthly_cap
                else "At cap" if count == monthly_cap
                else "Under cap"
            ),
        }
        for (eng_name, year_month), count in engineer_load.items()
    ]

    return assignments_df, pd.DataFrame(summary_records)


# Writing results back into the formatted master file

def rebuild_hayley_schedule(schedule, assignments_df, master_path, out_path):
    # The point of this function is to keep the master's colours, row highlights
    # and layout. We do that by copying the original file and only changing the
    # numbers in the month cells, rather than writing a fresh sheet from scratch.

    # Nothing was assigned, so just hand back the master untouched.
    if assignments_df.empty:
        shutil.copy(master_path, out_path)
        return schedule

    # Turn the assignments into a postcode-by-month table of job counts.
    pivot = assignments_df.pivot_table(
        index="PostcodeExample",
        columns="PlannedMonthName",
        values="Jobs_Assigned",
        aggfunc="sum",
        fill_value=0,
    )
    for m in MONTH_ORDER:
        if m not in pivot.columns:
            pivot[m] = 0
    pivot = pivot[MONTH_ORDER]

    # Open a copy of the original and overwrite only the month values.
    shutil.copy(master_path, out_path)
    wb = load_workbook(out_path)
    ws = wb.active

    month_col = {}
    postcode_col = None
    for cell in ws[1]:
        if cell.value in MONTH_ORDER:
            month_col[cell.value] = cell.column
        if str(cell.value).strip() == "Postcode":
            postcode_col = cell.column

    for row_idx in range(2, ws.max_row + 1):
        pc_cell = ws.cell(row=row_idx, column=postcode_col)
        pc = str(pc_cell.value).strip() if pc_cell.value is not None else ""
        if pc not in pivot.index:
            continue
        for m, col in month_col.items():
            val = pivot.loc[pc, m]
            cell = ws.cell(row=row_idx, column=col)
            # Empty months go back to blank so we don't litter the sheet with zeros.
            cell.value = int(val) if val and not pd.isna(val) else None

    wb.save(out_path)
    return schedule


# Main entry point

def main(
    monthly_cap: int = DEFAULT_MONTHLY_CAP,
    engineers_override_path: Optional[Path] = None,
    input_dir: Optional[Path] = None,
    output_dir: Optional[Path] = None,
):
    # Fall back to the default data folders if the caller didn't pass any in.
    input_dir = Path(input_dir) if input_dir else _DEFAULT_INPUT_DIR
    output_dir = Path(output_dir) if output_dir else _DEFAULT_OUTPUT_DIR

    job_path = input_dir / "job_data.xlsx"
    eng_path = engineers_override_path or (input_dir / "engineers.xlsx")
    sched_path = input_dir / "SERVICE_SCHEDULE_MASTER.xlsx"
    outcode_path = input_dir / "postcode-outcodes.csv"

    print("=== Service Scheduler - Preview (Steps 1-8) ===\n")

    print(f"Loading job data from:          {job_path}")
    jobs = pd.read_excel(job_path)

    print(f"Loading engineers from:         {eng_path}")
    engineers = pd.read_excel(eng_path)

    print(f"Loading schedule master from:   {sched_path}")
    schedule = pd.read_excel(sched_path)

    print(f"Loading postcode outcodes from: {outcode_path}")
    outcodes = pd.read_csv(outcode_path)

    print("\n--- Columns detected ---")
    print("Jobs columns:      ", jobs.columns.tolist())
    print("Engineers columns: ", engineers.columns.tolist())
    print("Schedule columns:  ", schedule.columns.tolist())
    print("Outcodes columns:  ", outcodes.columns.tolist())

    jobs_with_area = add_postal_area_to_jobs(jobs)
    engineers_with_area = add_postal_area_to_engineers(engineers)
    cycles_df = build_schedule_cycles(schedule)

    output_dir.mkdir(parents=True, exist_ok=True)
    cycles_df.to_excel(output_dir / "schedule_cycles_preview.xlsx", index=False)

    outcode_coords, area_centroids = build_geo_maps(outcodes)
    print(f"Outcode coords count: {len(outcode_coords)}")
    print(f"Area centroids count: {len(area_centroids)}")

    area_eng_map, all_engineers = build_area_engineer_distance_map(
        engineers_with_area, outcode_coords, area_centroids, cycles_df
    )

    assignments_df, summary_df = assign_jobs_to_engineers(
        cycles_df, jobs_with_area, area_eng_map, all_engineers, monthly_cap
    )

    assignments_df.to_excel(output_dir / "engineer_assignments_preview.xlsx", index=False)
    summary_df.to_excel(output_dir / "engineer_monthly_summary.xlsx", index=False)

    # Build the updated schedule from the original master so it keeps its formatting.
    rebuild_hayley_schedule(
        schedule,
        assignments_df,
        master_path=sched_path,
        out_path=output_dir / "SERVICE_SCHEDULE_UPDATED.xlsx",
    )

    print("\nScheduler complete.")


if __name__ == "__main__":
    main()